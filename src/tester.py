import os

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font

from .batch import BatchSender
from .logger import logger
from .session import Session
from .testcase import BatchTester

test_report_dir = os.getenv("TEST_REPORT_DIR", "./test_reports")
if not os.path.exists(test_report_dir):
    os.makedirs(test_report_dir)


class Tester(object):
    def __init__(self, label,
                 user_info_queue,
                 url: str,
                 req_wrapper: callable,
                 title: str,
                 thread_cnt: int = 50,
                 start_func: callable = None,
                 session_update_func: callable = None,
                 stop_func: callable = None,
                 ):
        self.label = label
        self.user_info_queue = user_info_queue
        self.url = url
        self.req_wrapper = req_wrapper
        self.start_func = start_func
        self.session_update_func = session_update_func
        self.stop_func = stop_func
        self.thread_cnt = thread_cnt
        self.title = title
        self.session_cnt_to_check = self.user_info_queue.qsize()

        # 内部的变量
        self.batch_sender = BatchSender(
            label=label,
            url=self.url,
            thread_cnt=self.thread_cnt
        )

        self.bt: BatchTester = None

    def run(self, test_cases, only_check=False, session_cnt_to_check: int = 0, clear_session=False):
        if clear_session and only_check:
            raise ValueError("clear_session and only_check cannot be True at the same time")

        if clear_session:
            Session.clear_sessions(self.label)
            logger.info("清除会话数据成功")

        if not only_check:
            self.batch_sender.run(self.user_info_queue, self.req_wrapper,
                                  start_func=self.start_func,
                                  session_update_func=self.session_update_func,
                                  stop_func=self.stop_func)
            logger.info("发送请求完成")

        if session_cnt_to_check and session_cnt_to_check > self.session_cnt_to_check:
            self.session_cnt_to_check = session_cnt_to_check
            logger.info(f"检查的数量设置为{self.session_cnt_to_check}(限定会话数量)")
        else:
            logger.info(f"检查的数量设置为{self.session_cnt_to_check}(队列用户数量)")

        # 加载会话结果
        session_list = Session.load_sessions(self.label, n=self.session_cnt_to_check)
        if len(session_list) < session_cnt_to_check:
            raise ValueError(f"No enough sessions data found. expect[{session_cnt_to_check}], got[{len(session_list)}]")

        # 设置测试用例
        self.bt = BatchTester(self.title, session_list)
        logger.info("开始校验测试用例")

        # 进行校验
        self.bt.check(test_cases)
        logger.info("校验完成")

        # 产生报告
        logger.info("开始生成报告")
        self.gen_report()
        return session_list

    def gen_report(self):
        # 解析数据
        parsed_data = []
        ext_report = []
        for report in self.bt.report_list:
            report.summary()
            # 创建一个包含字典数据的列表
            parsed_data.append({
                "功能点": report.name,
                "预期结果": report.expectation,
                "最终结果": report.result,
                "异常说明": report.bad_case
            })

            if report.ext_report:
                ext_report.append((report.name, report.ext_report))

        # 保存为Excel文件
        output_file = os.path.join(test_report_dir, f"测试报告-{self.title}.xlsx")

        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            # 汇总信息
            df = pd.DataFrame(parsed_data)
            df.to_excel(writer, sheet_name="测试汇总", index=False)
            # 报告详情
            for name, ext_report in ext_report:
                ext_df = pd.DataFrame(ext_report)
                ext_df.to_excel(writer, sheet_name=name, index=False)

        # 加载生成的Excel文件
        wb = load_workbook(output_file)
        ws = wb.active

        # 设置列宽
        column_widths = {
            'A': 40,  # 功能点
            'B': 60,  # 预期结果
            'C': 15,  # 最终结果
            'D': 55  # 异常说明
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        font = Font(size=16)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                if cell.row == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=16, bold=True)
                elif cell.column_letter in ['A', 'C']:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif cell.column_letter in ['B', 'D']:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 自动调整行高
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    max_height = max(max_height, len(lines))
            if max_height > 0:
                ws.row_dimensions[row[0].row].height = max_height * 20  # 20是一个经验值，可以根据需要调整

        # 保存调整后的Excel文件
        wb.save(output_file)

        logger.info(f"数据已成功保存到 {output_file}")
