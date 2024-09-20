import os

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font

from client import BatchTester, BatchSender, load_sessions

test_report_dir = os.getenv("TEST_REPORT_DIR", "./test_reports")
if not os.path.exists(test_report_dir):
    os.makedirs(test_report_dir)


class Tester(object):
    def __init__(self, env, cred_id,
                 user_info_queue,
                 url: str,
                 req_wrapper: callable,
                 title: str,
                 thread_cnt: int = 50,
                 session_update_func: callable = None,
                 stop_func: callable = None,
                 ):
        self.env = env
        self.cred_id = cred_id
        self.user_info_queue = user_info_queue
        self.url = url
        self.req_wrapper = req_wrapper
        self.session_update_func = session_update_func
        self.stop_func = stop_func
        self.thread_cnt = thread_cnt
        self.title = title
        self.session_cnt_to_check = self.user_info_queue.qsize()

        # 内部的变量
        self.batch_sender = BatchSender(
            env=self.env,
            cred_id=self.cred_id,
            url=self.url,
            thread_cnt=self.thread_cnt
        )

        self.bt: BatchTester = None

    def run(self, test_cases, only_check=False, session_cnt_to_check: int = 0):
        if not only_check:
            self.batch_sender.run(self.user_info_queue, self.req_wrapper, self.session_update_func, self.stop_func)

        if session_cnt_to_check:
            if session_cnt_to_check > self.session_cnt_to_check:
                self.session_cnt_to_check = session_cnt_to_check

        # 加载会话结果
        session_list = load_sessions(env=self.env, cred_id=self.cred_id, n=self.session_cnt_to_check)

        # 设置测试用例
        self.bt = BatchTester(self.title, session_list)

        # 进行校验
        self.bt.check(test_cases)

        # 产生报告
        self.gen_report()
        return session_list

    def gen_report(self):
        # 解析数据
        parsed_data = []
        for report in self.bt.report_list:
            data_dict = report.summary_dict()
            # 创建一个包含字典数据的列表
            parsed_data.append({
                "功能点": data_dict['name'],
                "预期结果": data_dict['expectation'],
                "最终结果": data_dict['result'],
                "异常说明": data_dict['bad_case']
            })

        # 创建DataFrame
        df = pd.DataFrame(parsed_data)

        # 保存为Excel文件
        output_file = os.path.join(test_report_dir, f"测试报告-{self.title}.xlsx")
        df.to_excel(output_file, index=False)

        # 加载生成的Excel文件
        wb = load_workbook(output_file)
        ws = wb.active

        # 设置列宽
        column_widths = {
            'A': 40,  # 功能点
            'B': 60,  # 预期结果
            'C': 15,  # 最终结果
            'D': 55  # 异常说明
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        font = Font(size=16)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                if cell.row == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=16, bold=True)
                elif cell.column_letter in ['A', 'C']:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif cell.column_letter in ['B', 'D']:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 自动调整行高
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    max_height = max(max_height, len(lines))
            if max_height > 0:
                ws.row_dimensions[row[0].row].height = max_height * 20  # 20是一个经验值，可以根据需要调整

        # 保存调整后的Excel文件
        wb.save(output_file)

        print(f"数据已成功保存到 {output_file}")
