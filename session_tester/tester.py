import os
from typing import List

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font

from .logger import logger
from .session import update_test_session_dir
from .test_suite import TestSuite
from .testcase import Report

test_report_dir = os.getenv("TEST_REPORT_DIR", "./test_reports")
if not os.path.exists(test_report_dir):
    os.makedirs(test_report_dir)


class Tester:
    RUN_MODE_NEW = 0
    RUN_MODE_CHECK = 1
    RUN_MODE_BENCHMARK = 2

    def __init__(self,
                 name: str,
                 test_suites: List[TestSuite]):
        self.name = name
        self.test_suites = test_suites
        # Check no duplicate names
        names = [ts.name for ts in test_suites]
        if len(names) != len(set(names)):
            raise ValueError("Duplicate test suite names")

        for test_suite in test_suites:
            names = [tc.name for tc in test_suite.check_cases()]
            if len(names) != len(set(names)):
                raise ValueError(f"Duplicate test case names in suite {test_suite.name}")
        update_test_session_dir(self.name)

    def run(self, mode=RUN_MODE_NEW, thread_cnt=50):
        if mode not in [self.RUN_MODE_NEW, self.RUN_MODE_CHECK, self.RUN_MODE_BENCHMARK]:
            raise ValueError(f"Invalid tester run mode: {mode}")

        if mode == Tester.RUN_MODE_NEW:
            for test_suite in self.test_suites:
                test_suite.clear_sessions()
            logger.info("清除会话数据成功")
            for test_suite in self.test_suites:
                test_suite.do_send(thread_cnt=thread_cnt)
            logger.info("发送请求完成")
        elif mode == Tester.RUN_MODE_BENCHMARK:
            logger.info("启动压力测试")
            for test_suite in self.test_suites:
                result = test_suite.do_send(thread_cnt=thread_cnt, no_dump=True)
                result.report()
            logger.info("压测请求完成")

        # 只有新模式和校验模式下才会执行校验
        if mode in [Tester.RUN_MODE_NEW, Tester.RUN_MODE_CHECK]:
            for test_suite in self.test_suites:
                test_suite.check()
                logger.info(f"{test_suite.name}校验完成")

            # 保存为Excel文件
            output_file = os.path.join(test_report_dir, f"测试报告-{self.name}.xlsx")
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                self.gen_summary(writer)
                self.gen_detail_report(writer)
            self.format()

    def gen_summary(self, writer):
        # 解析数据
        parsed_data = []
        reports: List[Report]
        for test_suite in self.test_suites:
            reports = test_suite.report_list
            for report in reports:
                report.summary()
                # 创建一个包含字典数据的列表
                parsed_data.append({
                    "功能模块": test_suite.name,
                    "功能点": report.name,
                    "预期结果": report.expectation,
                    "最终结果": report.result,
                    "通过": report.passed_case_count,
                    "未通过": report.not_passed_case_count,
                    "未覆盖": report.uncover_case_count,
                    "网络错误": report.finished_with_err_count,
                    "异常说明": report.bad_case
                })

            # 保存为Excel文件
            output_file = os.path.join(test_report_dir, f"测试报告-{self.name}.xlsx")

            # 汇总信息
            df = pd.DataFrame(parsed_data)
            df.to_excel(writer, sheet_name="测试汇总", index=False)

        logger.info(f"汇总报告-已成功保存到 {output_file}")

    def gen_detail_report(self, writer):

        k = set()
        dup_test_case_name_set = set()
        for test_suite in self.test_suites:
            reports = test_suite.report_list
            for report in reports:
                report.summary()
                if not report.ext_report:
                    continue
                if report.name in k:
                    dup_test_case_name_set.add(report.name)
                k.add(report.name)

        for test_suite in self.test_suites:
            reports = test_suite.report_list
            for report in reports:
                report.summary()
                if not report.ext_report:
                    continue

                # 保存为Excel文件
                ext_df = pd.DataFrame(report.ext_report)
                if report.name in dup_test_case_name_set:
                    sheet_name = f"{report.name}({test_suite.name})"
                else:
                    sheet_name = report.name

                ext_df.to_excel(writer, sheet_name=sheet_name, index=False)
                logger.info(f"详细数据-已成功保存到 表-{sheet_name}")

    def format(self):
        output_file = os.path.join(test_report_dir, f"测试报告-{self.name}.xlsx")
        # 加载生成的Excel文件
        wb = load_workbook(output_file)
        ws = wb.active

        # 设置列宽
        column_widths = {
            'A': 30,  # 功能模块
            'B': 40,  # 功能点
            'C': 60,  # 预期结果
            'D': 15,  # 最终结果
            'E': 8.5,  # 通过
            'F': 10,  # 未通过
            'G': 10,  # 未覆盖
            'H': 13,  # 网络错误
            'I': 55,  # 异常说明
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        font = Font(size=16)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                if cell.row == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=16, bold=True)
                elif cell.column_letter in ['C', 'I']:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:  # elif cell.column_letter in ['A', 'B', 'D', 'E', 'G']:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 合并A列连续相同的单元格
        start_row = 2
        end_row = ws.max_row
        current_value = None
        merge_start = start_row
        for row in range(start_row, end_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value != current_value:
                if merge_start < row - 1:
                    ws.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
                current_value = cell_value
                merge_start = row

        # 最后一组相同值的单元格合并
        if merge_start < end_row:
            ws.merge_cells(start_row=merge_start, start_column=1, end_row=end_row, end_column=1)

        # 自动调整行高
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    max_height = max(max_height, len(lines))
            if max_height > 0:
                ws.row_dimensions[row[0].row].height = max_height * 20

        # 保存调整后的Excel文件
        wb.save(output_file)
